import json
import boto3
from io import BytesIO
from dotenv import load_dotenv
import os
from openpyxl import load_workbook
from dateutil import parser
import re
from utils.module import check_waiver_codes as obtain_code_entries
from utils.module import convert_state_code_to_name as state_code_to_name

# Load environment variables from .env file
load_dotenv()
aws_access_key_id = os.getenv('AWS_ACCESS_KEY_ID')
aws_secret_access_key = os.getenv('AWS_SECRET_ACCESS_KEY')
aws_region = os.getenv('AWS_REGION')

# Initialize boto3 S3 client with credentials
s3 = boto3.client(
    's3',
    aws_access_key_id=aws_access_key_id,
    aws_secret_access_key=aws_secret_access_key,
    region_name=aws_region
)

# Defining strings to be removed from objects
string_remove = {
    "location": "This certificate is issued for the operations specifically described hereinafter. No person shall conduct any operation pursuant to the"
}

# Used to create unformatted address
def strip_and_merge(strings_list):
    # Strip whitespace from each string in the list
    stripped_strings = [s.strip() for s in strings_list]
    # Merge the stripped strings with '/'
    merged_string = '/'.join(stripped_strings)
    return merged_string

# Extract needed information
def extract_final_info(blocks, key):
    info = {
        "Issued To": "",
        "Responsible Person": "",
        "Address": "",
        "Street Name and Number": "",
        "City": "",
        "State": "",
        "Zip Code": "",
        "Waiver Number": "",
        "Operations Authorized": "",
        "List of Waived Regulations": "",
        "Effective Date": "",
        "Expire Date": "",
        "Waiver URL": ""
    }

    capture_next = False
    address_line_count = 0

    info["Waiver URL"] = key.replace("json", "pdf").replace("waivers-pdf", "https://www.faa.gov/sites/faa.gov/files/")

    for block in blocks:
        if block['BlockType'] == 'LINE' and block['Page'] == 1:
            text = block['Text']

            if "ISSUED TO" in text:
                capture_next = "Issued To"
            elif "ADDRESS" in text:
                capture_next = "Address"
                address_line = [""] * 3
                address_line_count = 0
            elif "Responsible Person:" in text:
                info["Responsible Person"] = text.split(":", 1)[1].strip()
            elif "Responsible Party:" in text:
                info["Responsible Person"] = text.split(":", 1)[1].strip()
            elif "Waiver Number:" in text:
                info["Waiver Number"] = text.split(":", 1)[1].strip()
            elif "OPERATIONS AUTHORIZED" in text:
                capture_next = "Operations Authorized"
                info[capture_next] = ""  # Initialize as empty to append lines
            elif "LIST OF WAIVED REGULATIONS BY SECTION AND TITLE" in text:
                capture_next = "List of Waived Regulations"
                info[capture_next] = ""  # Initialize as empty to append lines
            elif "effective from" in text.lower():
                date_parts = text.split("effective from", 1)[1].split(" to ", 1)
                if len(date_parts) == 2:
                    start_date = parser.parse(date_parts[0].strip())
                    end_date = parser.parse(date_parts[1].strip().split(",")[0].strip())
                    info["Effective Date"] = start_date.strftime('%m/%d/%Y')
                    info["Expire Date"] = end_date.strftime('%m/%d/%Y')

            elif capture_next:
                if capture_next == "Address":
                    address_line[address_line_count] += (text + " ") if address_line_count < 3 else ""
                    address_line_count += 1
                    if address_line_count == 3:
                        info[capture_next] = strip_and_merge(address_line)
                        capture_next = False
                        address_components = info["Address"]  # get_address(info["Address"])
                        if address_components:
                            if len(address_line[2]) > 40:  # Case when we have three-row address
                                street_and_number, rest = address_components.split("/", 1)
                                cleaned_street_and_number = re.sub(r'[.,/]', '', street_and_number)
                                # Replace multiple spaces with a single space
                                info["Street Name and Number"] = re.sub(r'\s+', ' ', cleaned_street_and_number)
                                info["City"], rest = rest.split(",", 1)
                                state_zip = re.search(r'([A-Z]{2}) (\d{5})', rest)

                                if state_zip:
                                    info["State"] = state_code_to_name(state_zip.group(1))
                                    info["Zip Code"] = state_zip.group(2)
                                else:
                                    state = zip_code = None
                            else:
                                street_and_number = address_line[0] + '/' + address_line[1]
                                cleaned_street_and_number = re.sub(r'[.,/]', '', street_and_number)
                                # Replace multiple spaces with a single space
                                info["Street Name and Number"] = re.sub(r'\s+', ' ', cleaned_street_and_number)

                                city_state_zip = address_line[2]
                                info["City"], rest = city_state_zip.split(",", 1)
                                state_zip = re.search(r'([A-Z]{2}) (\d{5})', rest.strip())
                                if state_zip:
                                    info["State"] = state_code_to_name(state_zip.group(1))
                                    info["Zip Code"] = state_zip.group(2)
                                else:
                                    info["State"] = None
                                    info["Zip Code"] = None
                        else:
                            # Handle the case where address_components is None
                            info["Street Name and Number"] = "Unknown"
                            info["City"] = "Unknown"
                            info["State"] = "Unknown"
                            info["Zip Code"] = "Unknown"

                elif capture_next == "List of Waived Regulations":
                    # Append each new line to the List of Waived Regulations
                    info[capture_next] += text + " "
                    # Continue capturing until a new section starts
                    if any (keyword in text for keyword in ["STANDARD PROVISIONS"]):
                        capture_next = False
                elif capture_next == "Operations Authorized":
                    # Append each new line to the List of Waived Regulations
                    info[capture_next] += text + " "
                    # Continue capturing until a new section starts
                    if any (keyword in text for keyword in ["LIST OF WAIVED REGULATIONS BY SECTION AND TITLE"]):
                        capture_next = False
                else:
                    info[capture_next] = text
                    capture_next = False

    for key in info.keys():
        info[key] = info[key].strip()

    # Remove 'STANDARD PROVISIONS' from the final output if it was appended
    if "STANDARD PROVISIONS" in info["List of Waived Regulations"]:
        info["List of Waived Regulations"] = info["List of Waived Regulations"].replace("STANDARD PROVISIONS",
                                                                                        "").strip()

    # Remove 'LIST OF WAIVED REGULATIONS BY SECTION AND TITLE' from the final output if it was appended
    if "LIST OF WAIVED REGULATIONS BY SECTION AND TITLE" in info["Operations Authorized"]:
        info["Operations Authorized"] = info["Operations Authorized"].replace(
            "LIST OF WAIVED REGULATIONS BY SECTION AND TITLE", "").strip()

    # Remove redundant string from ["Address"]
    if string_remove["location"] in info["Address"]:
        info["Address"] = info["Address"].replace(
            string_remove["location"], "").strip()

    return info

# Find the first empty row in a sheet
def find_first_empty_row(sheet):
    for row in range(2, sheet.max_row + 2):
        if not sheet.cell(row=row, column=1).value:
            return row
    return sheet.max_row + 1

# Remove external links in the workbook
def remove_external_links(workbook):
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.value and cell.value.startswith('['):
                    cell.value = None

# Find the next available Company ID
def get_next_company_id(locations_sheet):
    company_ids = [int(row[1][1:]) for row in locations_sheet.iter_rows(min_row=2, values_only=True) if row and row[1] and row[1].startswith('C')]
    return f"C{(max(company_ids) + 1) if company_ids else 1}"

# Find the next available Operator ID
def get_next_operator_id(locations_sheet):
    operator_ids = [row[0] for row in locations_sheet.iter_rows(min_row=2, values_only=True) if row and row[0] is not None]
    return (max(operator_ids) + 1) if operator_ids else 1

# S3 storage values
bucket_name = 'auvsi-uav-waivers'
input_prefix = 'waivers-json/'
output_file_key = 'waivers_info.xlsx'

# Read the existing Excel file from S3
response = s3.get_object(Bucket=bucket_name, Key=output_file_key)
existing_file = response['Body'].read()

wb = load_workbook(filename=BytesIO(existing_file))
remove_external_links(wb)
waiver_data_sheet = wb["Waiver Data"]
locations_sheet = wb["Locations"]

# List objects in the specified S3 bucket and prefix
response = s3.list_objects_v2(Bucket=bucket_name, Prefix=input_prefix)

# Iterate through the files in the bucket
for obj in response.get('Contents', []):
    if obj['Key'].endswith('.json'):
        # Read the JSON file from S3
        json_file = s3.get_object(Bucket=bucket_name, Key=obj['Key'])
        textract_response = json.loads(json_file['Body'].read())
        final_extracted_info = extract_final_info(textract_response['Blocks'], obj['Key'])

        responsible_person = final_extracted_info["Responsible Person"].strip()
        address_street = final_extracted_info["Street Name and Number"].strip()
        address_city = final_extracted_info["City"].strip()
        address_state = final_extracted_info["State"].strip()
        address_zip = final_extracted_info["Zip Code"].strip()

        # Initialize operator_id, company_id, and full_operator_id with default values
        operator_id = None
        company_id = None
        full_operator_id = None

        # Find all rows with matching Responsible Person
        matching_rows = []
        for row in locations_sheet.iter_rows(min_row=2, values_only=True):
            if row and len(row) > 3 and row[3] and row[3].strip() == responsible_person:
                matching_rows.append(row)

        # Check each matching row for address match
        address_match_found = False
        for row in matching_rows:
            if row[4].strip() == address_street:  # Assuming address is in the 5th column (index 4)
                address_match_found = True
                if row[11] and row[11].strip().replace(".", "").replace(",", "") == final_extracted_info["Issued To"].strip().replace(".", "").replace(",", ""):
                    break
                else:
                    # Different company, new Company ID needed
                    if responsible_person == final_extracted_info["Issued To"]:
                        company_id = "INDIVIDUAL"
                    else:
                        company_id = get_next_company_id(locations_sheet)
                    full_operator_id = f"{row[0]}-{company_id}"
                    break

        if not matching_rows:
            # No matching Responsible Person found, assign new Operator ID and Company ID
            operator_id = get_next_operator_id(locations_sheet)
            if responsible_person == final_extracted_info["Issued To"]:
                company_id = "INDIVIDUAL"
            else:
                company_id = get_next_company_id(locations_sheet)
            full_operator_id = f"{operator_id}-{company_id}"

        elif not address_match_found:
            # Matching Responsible Person found but no matching address
            operator_id = matching_rows[0][0]  # Use the operator ID from the first match
            if responsible_person != final_extracted_info["Issued To"]:
                company_id = get_next_company_id(locations_sheet)
            else:
                company_id = "INDIVIDUAL"
            full_operator_id = f"{operator_id}-{company_id}"

        elif address_match_found:
            operator_id = matching_rows[0][0]  # Use the operator ID from the first match
            if responsible_person != final_extracted_info["Issued To"]:
                company_id = get_next_company_id(locations_sheet)
            else:
                company_id = "INDIVIDUAL"
            full_operator_id = f"{operator_id}-{company_id}"

        # Skip adding new entry if an exact match is found
        if not address_match_found:
            new_location = [
                operator_id, company_id, full_operator_id, responsible_person,
                address_street, address_city, address_state, address_zip,
                "", "", "", "" if company_id == "INDIVIDUAL" else final_extracted_info["Issued To"]
            ]
            first_empty_row = find_first_empty_row(locations_sheet)
            for col, value in enumerate(new_location, start=1):
                locations_sheet.cell(row=first_empty_row, column=col, value=value)

        # Add entry to Waiver Data sheet
        effective_date = final_extracted_info["Effective Date"]
        expire_date = final_extracted_info["Expire Date"]
        waiver_url = final_extracted_info["Waiver URL"]
        waived_regulations = obtain_code_entries(final_extracted_info["List of Waived Regulations"])

        new_waiver_data = [
            operator_id, company_id, full_operator_id, effective_date,
            expire_date, final_extracted_info["Waiver Number"],
            waiver_url,
            waived_regulations["Daylight Operations"],
            waived_regulations["VLOS Operations"],
            waived_regulations["Visual Observer"],
            waived_regulations["Multiple UAS"],
            waived_regulations["Over People"],
            waived_regulations["Operation in Certain Airspace"],
            waived_regulations["Operating Limitations (a)"],
            waived_regulations["Operating Limitations (b, c, d)"],
            waived_regulations["Moving Vehicle or Aircraft"],
            waived_regulations["Over Moving Vehicles"],
            final_extracted_info["Operations Authorized"]
        ]
        first_empty_row = find_first_empty_row(waiver_data_sheet)
        for col, value in enumerate(new_waiver_data, start=1):
            waiver_data_sheet.cell(row=first_empty_row, column=col, value=value)

# Save the updated Excel file to a BytesIO object
output = BytesIO()
wb.save(output)
output.seek(0)

# Upload the updated file back to S3
s3.put_object(Bucket=bucket_name, Key=output_file_key, Body=output)

print(f"Data successfully appended to {output_file_key}")
